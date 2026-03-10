# -*- coding: utf-8 -*-
"""
Excel 智能整理工具

使用说明
1. 安装依赖：pip install flask openpyxl pandas requests
2. 运行程序：python excel_tidy.py
3. 打开浏览器访问 http://127.0.0.1:5678

说明
- 支持 .xlsx / .xls；读取 .xls 时依赖 pandas 的 xlrd 引擎
- 优先使用本地 SmartTidy，引擎失败时可回退 LLM
- 前后端均使用 UTF-8 编码，避免中文乱码
"""

from __future__ import annotations

import json
import os
import re
import sys
import traceback
import uuid
import webbrowser
from io import BytesIO
from threading import Timer
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
from flask import Flask, Response, jsonify, request, send_file
from openpyxl.utils.dataframe import dataframe_to_rows

# ==============================
# 全局状态
# ==============================

data_store: Dict[str, Dict[str, Any]] = {}
raw_file_store: Dict[str, Dict[str, Any]] = {}
PROGRAM_DIR = os.path.dirname(os.path.abspath(sys.argv[0]))


# ==============================
# 通用工具
# ==============================

def sanitize_dataframe(df: Optional[pd.DataFrame]) -> pd.DataFrame:
    """清洗 DataFrame，处理 NaN/Infinity 并统一可序列化值。"""
    if df is None:
        return pd.DataFrame()
    safe = df.copy()
    safe = safe.replace([float("inf"), float("-inf")], pd.NA)
    safe = safe.where(pd.notna(safe), None)
    safe.columns = [str(col).strip() for col in safe.columns]
    return safe


def _json_safe_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        if isinstance(value, float) and (pd.isna(value) or value in (float("inf"), float("-inf"))):
            return None
        return value
    return str(value)


def dataframe_preview(df: Optional[pd.DataFrame], max_rows: int = 80) -> List[Dict[str, Any]]:
    if df is None or df.empty:
        return []
    preview_df = sanitize_dataframe(df.head(max_rows))
    records: List[Dict[str, Any]] = []
    for row in preview_df.to_dict(orient="records"):
        records.append({str(k): _json_safe_value(v) for k, v in row.items()})
    return records


def _list_sheet_names(file_content: bytes, ext: str) -> List[str]:
    if ext == ".xlsx":
        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    xls = pd.ExcelFile(BytesIO(file_content))
    return list(xls.sheet_names)


def _read_sheet_dataframe(file_content: bytes, ext: str, sheet_name: Optional[str]) -> pd.DataFrame:
    stream = BytesIO(file_content)
    kwargs: Dict[str, Any] = {"sheet_name": sheet_name, "dtype": object}
    if ext == ".xlsx":
        kwargs["engine"] = "openpyxl"
    df = pd.read_excel(stream, **kwargs)
    if not isinstance(df, pd.DataFrame):
        raise ValueError("读取结果不是有效的 DataFrame")
    return sanitize_dataframe(df)


def _convert_xls_to_xlsx_bytes(file_content: bytes) -> bytes:
    """将 .xls 转为 .xlsx，便于统一使用 openpyxl 处理。"""
    try:
        all_sheets = pd.read_excel(BytesIO(file_content), sheet_name=None, header=None, dtype=object)
    except Exception as exc:
        raise ValueError("读取 .xls 失败，请安装 xlrd 或改用 .xlsx 文件") from exc
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for idx, (sheet_name, df) in enumerate(all_sheets.items()):
        safe_name = str(sheet_name).strip() or f"Sheet{idx + 1}"
        ws = wb.create_sheet(title=safe_name[:31])
        safe_df = sanitize_dataframe(df)
        for row in dataframe_to_rows(safe_df, index=False, header=False):
            ws.append(list(row))
    output = BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)
    return output.read()


# ==============================
# smart_tidy.py 逻辑
# ==============================
AVERAGE_KEYWORDS = ('平均', '均值', '平均值', 'average', 'avg', 'mean')

HEADER_STOP_WORDS = {'处理', '重复', '平均', '均值', '平均值'}

class TableSpec(dict):
    """Small helper type to clarify spec fields."""

def _clean_text(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None

def _build_merged_lookup(ws):
    merged_values = {}
    merged_spans = {}
    for merged in ws.merged_cells.ranges:
        value = ws.cell(merged.min_row, merged.min_col).value
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                merged_values[row, col] = value
                merged_spans[row, col] = (merged.min_row, merged.max_row, merged.min_col, merged.max_col)
    return (merged_values, merged_spans)

def _is_average_label(text: Optional[str]) -> bool:
    if not text:
        return False
    lower = text.lower()
    return any((keyword in lower for keyword in AVERAGE_KEYWORDS))

def _ranges_overlap(a_start: int, a_end: int, b_start: int, b_end: int) -> bool:
    return a_start <= b_end and b_start <= a_end

def _to_digit(token: str) -> Optional[int]:
    if not token:
        return None
    match = re.search('\\d+', token)
    if match:
        return int(match.group(0))
    mapping = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '七': 7, '八': 8, '九': 9, '十': 10}
    return mapping.get(token)

def _parse_repeat_label(text: Optional[str]) -> Optional[str]:
    if not text:
        return None
    compact = re.sub('\\s+', '', text)
    if _is_average_label(compact):
        return '__AVG__'
    match = re.search('重复([一二三四五六七八九十\\d]+)', compact)
    if match:
        number = _to_digit(match.group(1))
        if number is not None:
            return f'重复{number}'
    if compact in {'1', '2', '3', '4', '5'}:
        return f'重复{compact}'
    return None

def _normalize_metric_name(metric: Optional[str], table_title: str) -> str:
    text = _clean_text(metric)
    if not text or text == '处理':
        return '值'
    if text == table_title:
        return '值'
    return text

def _coerce_value(value):
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value

def _first_non_empty(values):
    for value in values:
        if value is None:
            continue
        if isinstance(value, float) and pd.isna(value):
            continue
        return value
    return None

def _pick_treatment_from_row(row: int, key_col_start: int, key_col_end: int, get_val) -> Optional[str]:
    key_values = []
    for col in range(key_col_start, key_col_end + 1):
        value = _clean_text(get_val(row, col))
        if value:
            key_values.append(value)
    if not key_values:
        return None
    candidate = key_values[-1]
    if candidate in HEADER_STOP_WORDS:
        return None
    if _is_average_label(candidate):
        return None
    if candidate.startswith('重复'):
        return None
    return candidate

def _is_title_candidate(value: Optional[str]) -> bool:
    text = _clean_text(value)
    if not text:
        return False
    if text in HEADER_STOP_WORDS:
        return False
    if _is_average_label(text):
        return False
    metric_prefixes = ('茎', '叶', '穗', '总重')
    if text.startswith(metric_prefixes):
        return False
    return True

def _find_key_anchor_for_title(ws, get_val, merged_spans, merge_range) -> Optional[Tuple[int, int, int]]:
    title_row = merge_range.min_row
    col_start = max(1, merge_range.min_col - 2)
    col_end = merge_range.max_col
    for row in range(title_row, min(ws.max_row, title_row + 2) + 1):
        for col in range(col_start, col_end + 1):
            value = _clean_text(get_val(row, col))
            if value != '处理':
                continue
            span = merged_spans.get((row, col))
            if span:
                _, _, key_start, key_end = span
            else:
                key_start = col
                key_end = col
            return (row, key_start, key_end)
    return None

def _expand_value_col_end(ws, key_row: int, key_col_end: int, current_end: int, get_val) -> int:
    """
    Expand value-column boundary using merged metric headers near key/header rows.
    This is required for cases where a title merge is narrower than the real metric area.
    """
    expanded_end = current_end
    changed = True
    while changed:
        changed = False
        for merged in ws.merged_cells.ranges:
            row_span = merged.max_row - merged.min_row + 1
            col_span = merged.max_col - merged.min_col + 1
            if row_span != 1 or col_span < 2:
                continue
            if not key_row <= merged.min_row <= key_row + 1:
                continue
            if merged.min_col <= key_col_end:
                continue
            if merged.min_col > expanded_end + 2:
                continue
            value = _clean_text(get_val(merged.min_row, merged.min_col))
            if not value or value in HEADER_STOP_WORDS or _is_average_label(value):
                continue
            if merged.max_col > expanded_end:
                expanded_end = merged.max_col
                changed = True
    return expanded_end

def _detect_table_specs(ws, get_val, merged_spans) -> List[TableSpec]:
    specs: List[TableSpec] = []
    merged_ranges = sorted(ws.merged_cells.ranges, key=lambda item: (item.min_row, item.min_col, item.max_col))
    for merged in merged_ranges:
        row_span = merged.max_row - merged.min_row + 1
        col_span = merged.max_col - merged.min_col + 1
        if row_span != 1 or col_span < 4:
            continue
        title = _clean_text(get_val(merged.min_row, merged.min_col))
        if not _is_title_candidate(title):
            continue
        key_anchor = _find_key_anchor_for_title(ws, get_val, merged_spans, merged)
        if key_anchor is None:
            continue
        key_row, key_col_start, key_col_end = key_anchor
        value_col_start = max(key_col_end + 1, merged.min_col)
        value_col_end = _expand_value_col_end(ws=ws, key_row=key_row, key_col_end=key_col_end, current_end=merged.max_col, get_val=get_val)
        if value_col_start > value_col_end:
            continue
        specs.append(TableSpec(title=title, title_row=merged.min_row, key_row=key_row, key_col_start=key_col_start, key_col_end=key_col_end, value_col_start=value_col_start, value_col_end=value_col_end, table_col_start=min(key_col_start, merged.min_col), table_col_end=max(value_col_end, key_col_end)))
    deduped: List[TableSpec] = []
    seen = set()
    for spec in specs:
        key = (spec['title_row'], spec['table_col_start'], spec['table_col_end'], spec['title'])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(spec)
    deduped.sort(key=lambda item: (item['title_row'], item['table_col_start']))
    for index, spec in enumerate(deduped):
        boundary_end = ws.max_row
        for candidate in deduped[index + 1:]:
            if not _ranges_overlap(spec['table_col_start'], spec['table_col_end'], candidate['table_col_start'], candidate['table_col_end']):
                continue
            boundary_end = candidate['title_row'] - 1
            break
        end_row = _find_table_end_row(spec, boundary_end, get_val)
        spec['end_row'] = max(end_row, spec['key_row'] + 1)
    return deduped

def _find_table_end_row(spec: TableSpec, boundary_end: int, get_val) -> int:
    start_row = spec['key_row'] + 1
    last_data_row = start_row - 1
    empty_streak = 0
    for row in range(start_row, boundary_end + 1):
        treatment = _pick_treatment_from_row(row, spec['key_col_start'], spec['key_col_end'], get_val)
        if treatment:
            last_data_row = row
            empty_streak = 0
        elif last_data_row >= start_row:
            empty_streak += 1
            if empty_streak >= 2:
                break
    return max(last_data_row, start_row)

def _extract_subtable_dataframe(ws, get_val, spec: TableSpec, drop_avg: bool) -> pd.DataFrame:
    key_row = spec['key_row']
    sub_header_row = key_row + 1 if key_row + 1 <= spec['end_row'] else key_row
    column_defs = []
    has_repeat_sub_headers = False
    for col in range(spec['value_col_start'], spec['value_col_end'] + 1):
        top_label = _clean_text(get_val(key_row, col))
        sub_label = _clean_text(get_val(sub_header_row, col))
        repeat_label = _parse_repeat_label(sub_label)
        if repeat_label is not None:
            has_repeat_sub_headers = True
        column_defs.append({'col': col, 'metric': _normalize_metric_name(top_label, spec['title']), 'repeat': repeat_label, 'sub': sub_label})
    use_two_header_rows = has_repeat_sub_headers and sub_header_row > key_row
    data_start_row = key_row + (2 if use_two_header_rows else 1)
    repeat_metric_cols: Dict[str, Dict[str, List[int]]] = {}
    for definition in column_defs:
        repeat_label = definition['repeat']
        if repeat_label is None:
            if use_two_header_rows:
                continue
            repeat_label = '重复1'
        if drop_avg and repeat_label == '__AVG__':
            continue
        metric_name = definition['metric']
        if drop_avg and _is_average_label(metric_name):
            continue
        repeat_metric_cols.setdefault(repeat_label, {}).setdefault(metric_name, []).append(definition['col'])
    records = []
    for row in range(data_start_row, spec['end_row'] + 1):
        treatment = _pick_treatment_from_row(row, spec['key_col_start'], spec['key_col_end'], get_val)
        if not treatment:
            continue
        for repeat_label, metric_map in repeat_metric_cols.items():
            if repeat_label == '__AVG__':
                continue
            record = {'处理': treatment, '重复': repeat_label}
            non_null_values = 0
            for metric_name, 列 in metric_map.items():
                value = _first_non_empty([_coerce_value(get_val(row, col)) for col in 列])
                if value is not None:
                    non_null_values += 1
                record[f"{spec['title']}_{metric_name}"] = value
            if non_null_values > 0:
                records.append(record)
    if not records:
        return pd.DataFrame(columns=['处理', '重复'])
    df = pd.DataFrame(records)
    df = _coalesce_duplicate_rows(df)
    return df

def _coalesce_duplicate_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    key_columns = ['处理', '重复']
    metric_columns = [col for col in df.columns if col not in key_columns]
    if not metric_columns:
        return df.drop_duplicates(subset=key_columns).reset_index(drop=True)

    def pick_first_valid(series):
        for value in series:
            if value is None:
                continue
            if isinstance(value, float) and pd.isna(value):
                continue
            return value
        return None
    grouped = df.groupby(key_columns, sort=False, as_index=False)[metric_columns].agg(pick_first_valid).reset_index(drop=True)
    return grouped

def _merge_subtable_dataframes(frames: List[pd.DataFrame]) -> pd.DataFrame:
    valid_frames = [frame for frame in frames if frame is not None and (not frame.empty)]
    if not valid_frames:
        return pd.DataFrame(columns=['处理', '重复'])
    merged = valid_frames[0].copy()
    for frame in valid_frames[1:]:
        merged = pd.merge(merged, frame, on=['处理', '重复'], how='outer')
    merged['处理'] = merged['处理'].astype(str).str.strip()
    merged['重复'] = merged['重复'].astype(str).str.strip()
    ordered_cols = ['处理', '重复'] + [col for col in merged.columns if col not in {'处理', '重复'}]
    merged = merged[ordered_cols]
    merged = merged.drop_duplicates(subset=['处理', '重复']).reset_index(drop=True)
    return merged

def _serialize_preview(df: pd.DataFrame, preview_rows: int=20) -> List[Dict[str, object]]:
    if df.empty:
        return []
    preview_df = df.head(preview_rows).copy()
    preview_df = preview_df.where(pd.notna(preview_df), None)
    records = []
    for row in preview_df.to_dict(orient='records'):
        record = {}
        for key, value in row.items():
            if value is not None and (not isinstance(value, (str, int, float, bool))):
                record[key] = str(value)
            else:
                record[key] = value
        records.append(record)
    return records

def _parse_excel(file_content: bytes, sheet_name: Optional[str], drop_avg: bool):
    workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    try:
        if sheet_name and sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
        else:
            ws = workbook.active
        merged_values, merged_spans = _build_merged_lookup(ws)

        def get_val(row: int, col: int):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                return value
            return merged_values.get((row, col))
        specs = _detect_table_specs(ws, get_val, merged_spans)
        subtable_frames = []
        summaries = []
        for spec in specs:
            frame = _extract_subtable_dataframe(ws, get_val, spec, drop_avg=drop_avg)
            if frame is None or frame.empty:
                continue
            subtable_frames.append(frame)
            summaries.append({'title': spec['title'], 'rows': int(len(frame)), 'columns': int(len(frame.columns) - 2), 'row_range': [int(spec['title_row']), int(spec['end_row'])], 'column_range': [int(spec['table_col_start']), int(spec['table_col_end'])], 'factor_headers': ['处理', '重复']})
        merged_df = _merge_subtable_dataframes(subtable_frames)
        return {'sheet_name': ws.title, 'merged_count': len(ws.merged_cells.ranges), 'specs': specs, 'summaries': summaries, 'table_frames': subtable_frames, 'merged_df': merged_df}
    finally:
        workbook.close()

def scan_excel_structure(file_content, sheet_name=None):
    """Scan workbook and build a robust multi-subtable parsing plan."""
    if not file_content:
        return {'error': 'Excel 内容为空'}
    parsed = _parse_excel(file_content, sheet_name, drop_avg=True)
    merged_df = parsed['merged_df']
    if merged_df.empty:
        return {'error': '未检测到有效的子表数据'}
    title_parts = [item['title'] for item in parsed['summaries'] if item.get('title')]
    title = ' / '.join(title_parts) if title_parts else parsed['sheet_name']
    return {'title': title, 'sheet_name': parsed['sheet_name'], 'merged_count': parsed['merged_count'], 'sub_table_count': len(parsed['summaries']), 'sub_tables': parsed['summaries'], 'total_rows': int(len(merged_df)), 'total_cols': int(len(merged_df.columns)), 'preview': _serialize_preview(merged_df), '_file_content': file_content, '_sheet_name': sheet_name, '_drop_avg_df': merged_df, '_spec_count': len(parsed['specs'])}

def execute_smart_tidy(scan_result, options=None):
    """Execute transformation using scan result and return wide-format DataFrame."""
    if options is None:
        options = {}
    drop_avg = bool(options.get('drop_avg', True))
    if drop_avg and isinstance(scan_result.get('_drop_avg_df'), pd.DataFrame):
        return scan_result['_drop_avg_df'].copy()
    file_content = scan_result.get('_file_content')
    sheet_name = scan_result.get('_sheet_name')
    if not file_content:
        return pd.DataFrame(columns=['处理', '重复'])
    parsed = _parse_excel(file_content, sheet_name, drop_avg=drop_avg)
    return parsed['merged_df']

# ==============================
# llm_tidy.py 逻辑
# ==============================

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

DEFAULT_API_BASE = 'https://api.siliconflow.cn/v1'

DEFAULT_MODEL = 'deepseek-ai/DeepSeek-V3'


def _settings_file_path() -> str:
    if getattr(sys, 'frozen', False):
        settings_dir = os.path.dirname(os.path.abspath(sys.executable))
    else:
        settings_dir = PROGRAM_DIR
    return os.path.join(settings_dir, 'settings.json')


def _default_settings() -> Dict[str, str]:
    return {'api_key': '', 'model': DEFAULT_MODEL, 'api_base': DEFAULT_API_BASE}


def _load_settings() -> Dict[str, str]:
    settings = _default_settings()
    settings_path = _settings_file_path()
    if not os.path.exists(settings_path):
        return settings
    try:
        with open(settings_path, 'r', encoding='utf-8') as f:
            payload = json.load(f)
    except Exception:
        traceback.print_exc()
        return settings
    if isinstance(payload, dict):
        for key in ('api_key', 'model', 'api_base'):
            value = payload.get(key)
            if isinstance(value, str):
                settings[key] = value
    return settings


def _save_settings(settings: Dict[str, str]) -> None:
    payload = {key: (settings.get(key) or '') for key in ('api_key', 'model', 'api_base')}
    settings_path = _settings_file_path()
    tmp_path = os.path.join(PROGRAM_DIR, f'.settings.{uuid.uuid4().hex}.tmp')
    try:
        with open(tmp_path, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, settings_path)
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


app_settings: Dict[str, str] = _load_settings()

def read_excel_raw(file_content, sheet_name=None, max_rows=80, max_cols=25):
    """
    用 openpyxl 读取 Excel 原始内容（含合并单元格），
    返回文本描述让 LLM 能理解数据结构。
    """
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    total_rows = ws.max_row or 0
    total_cols = ws.max_column or 0
    merged_info = []
    merged_fill = {}
    for m in ws.merged_cells.ranges:
        val = ws.cell(m.min_row, m.min_col).value
        span_desc = f'R{m.min_row}C{m.min_col}:R{m.max_row}C{m.max_col}'
        merged_info.append(f'  {span_desc} => "{val}"  (跨{m.max_row - m.min_row + 1}行{m.max_col - m.min_col + 1}列)')
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                merged_fill[r, c] = val
    lines = []
    lines.append(f'=== Excel 工作表: {ws.title} ===')
    lines.append(f'总行数: {total_rows}, 总列数: {total_cols}')
    lines.append(f'')
    lines.append(f'合并单元格 ({len(merged_info)} 个):')
    for info in sorted(merged_info):
        lines.append(info)
    lines.append('')
    lines.append('全部单元格内容:')
    rows_to_show = min(max_rows, total_rows)
    for r in range(1, rows_to_show + 1):
        row_data = []
        for c in range(1, min(max_cols + 1, total_cols + 1)):
            val = ws.cell(r, c).value
            if val is None and (r, c) in merged_fill:
                val = merged_fill[r, c]
            if val is not None:
                sv = str(val)
                if len(sv) > 20:
                    sv = sv[:20] + '..'
                row_data.append(f'(R{r},C{c})={sv}')
        if row_data:
            lines.append(f"  Row{r}: {', '.join(row_data)}")
        else:
            lines.append(f'  Row{r}: (空行)')
    if total_rows > rows_to_show:
        lines.append(f'\n... 省略 Row{rows_to_show + 1}~Row{total_rows} ...')
    wb.close()
    return '\n'.join(lines)

def call_llm(api_key, prompt, system_prompt=None, api_base=None, model=None):
    """调用 LLM API（OpenAI 兼容格式）"""
    if not HAS_REQUESTS:
        raise RuntimeError('缺少 requests 库')
    base = (api_base or DEFAULT_API_BASE).rstrip('/')
    mdl = model or DEFAULT_MODEL
    url = f'{base}/chat/completions'
    messages = []
    if system_prompt:
        messages.append({'role': 'system', 'content': system_prompt})
    messages.append({'role': 'user', 'content': prompt})
    headers = {'Authorization': f'Bearer {api_key}', 'Content-Type': 'application/json'}
    payload = {'model': mdl, 'messages': messages, 'temperature': 0.05, 'max_tokens': 8192}
    resp = requests.post(url, headers=headers, json=payload, timeout=180)
    resp.raise_for_status()
    result = resp.json()
    return result['choices'][0]['message']['content']

SYSTEM_PROMPT = '你是一个专业的数据整理助手。用户会给你一个 Excel 工作表的原始单元格数据。\n\n## 你的任务\n\n生成 Python 代码将这个复杂的 Excel 转换为**标准分析用宽表格式**。\n\n## ⚠️ 绝对禁止写死行列号\n\n由于文本丢失了部分空间直觉，**你绝对不能在代码中硬编码任何具体的行列数字**（如 `min_row=16`, `range(3, 15)` 等）。\n你必须生成**动态搜索表块**的代码。\n\n## 数据布局与动态探测策略\n\n1. **结构特点**：一个表里散布着多个独立的数据表块。每个表块的**起点都是一个跨多列的标题合并单元格**（比如 "分蘖期干物质"、"孕穗-齐穗期积累" 等）。\n2. **如何动态定位（请在代码中实现类似逻辑）**：\n   - 第一步：遍历 `ws.merged_cells.ranges`，凡是值不为空（排除 None/处理/重复）、且位于某一行的起始位置、跨越多列的，都认为是"表块标题"。\n   - 第二步：根据标题合并单元格的 `min_row` 和 `min_col`。它的正下方（如 `min_row+1`, `min_row+2`）通常是列名行（指标名如"茎(g)"、处理、重复等）。\n   - 第三步：接着向下遍历获取该区域的独立数据，直到遇到空行跳出。\n   - 第四步：每个表块独立提取出一份 DataFrame。\n\n3. **左右交错问题**：不同表块（如干物质在左、积累量在右）存在于同一行中。靠着 `min_col` 与 `max_col` 范围来确保你只读取该块内的列。\n\n## 宽表格式要求\n\n- **每行** = 处理名 × 重复编号。不要把处理合并，每行代表一个生物学重复的唯一记录（例如处理A重复1、处理A重复2）。\n- **列** = 包含所有时期和指标。列名必须带表块标题前缀，如 `分蘖期干物质_茎(g)`、`分蘖-孕穗期积累_值`。\n- 移除所有名为"平均"、"均值"、"平均值"的列及其对应的数据。\n\n## 必须遵守的代码约束\n\n1. 手填合并单元格缓冲：\n```python\nmerged_vals = {}\nfor m in ws.merged_cells.ranges:\n    val = ws.cell(m.min_row, m.min_col).value\n    for r in range(m.min_row, m.max_row+1):\n        for c in range(m.min_col, m.max_col+1):\n            merged_vals[(r,c)] = val\n\ndef get_val(r, c):\n    v = ws.cell(r, c).value\n    return v if v is not None else merged_vals.get((r,c))\n```\n2. 表块横向合并(Merge)：把所有临时表块的数据，用 `pd.merge(..., on=[\'处理\',\'重复\'], how=\'outer\')` 合并。注意统一 `处理` 列转为 `str` 且去除空格。如果遇到缺少列名的空表头，请根据上下文推断或忽略。\n\n## 返回格式\n\n仅返回以下 JSON 对象（不要 ```json 包装，不要文字）。\n{\n  "description": "说明你对动态解析思路的规划",\n  "tables_found": [{"name": "此处列举你探测到的表名,非写死行列"}],\n  "code": "完整的 Python 代码（def transform(file_content, sheet_name=None): ...）"\n}\n\n- 只能用 openpyxl, pandas, io.BytesIO'

def analyze_and_transform(file_content, api_key, sheet_name=None, api_base=None, model=None, max_retries=3):
    """
    主入口：用 LLM 分析 Excel 结构并执行转换。
    如果生成的代码执行失败，会自动将错误反馈给 LLM 重试。

    Returns:
        dict: {
            'success': bool,
            'description': str,
            'tables_found': list,
            'result_df': DataFrame or None,
            'error': str or None,
            'llm_response': str
        }
    """
    raw_text = read_excel_raw(file_content, sheet_name, max_rows=80, max_cols=25)
    prompt = f'请分析以下 Excel 工作表的数据结构，并生成转换代码。\n\n{raw_text}\n\n请仔细观察：\n1. 有哪些独立的数据表块？各自在什么行列范围？\n2. 是否有不同列区域的表块共享同一行号（左右交错）？\n3. 每个表块的指标列有多少组？\n4. 请生成 transform 函数，将所有数据合并为一张宽表。\n\n返回 JSON 对象。'
    description = ''
    tables_found = []
    last_response = ''
    code = ''
    for attempt in range(max_retries + 1):
        try:
            print(f'[LLM Tidy] 第 {attempt + 1}/{max_retries + 1} 次调用 LLM...')
            llm_response = call_llm(api_key, prompt, SYSTEM_PROMPT, api_base, model)
            last_response = llm_response
            cleaned = llm_response.strip()
            if cleaned.startswith('```'):
                cleaned = re.sub('^```\\w*\\s*', '', cleaned)
                cleaned = re.sub('\\s*```$', '', cleaned)
            parsed = json.loads(cleaned)
            description = parsed.get('description', '')
            tables_found = parsed.get('tables_found', [])
            code = parsed.get('code', '')
            if not code:
                return {'success': False, 'description': description, 'tables_found': tables_found, 'result_df': None, 'error': 'LLM 未返回转换代码', 'llm_response': llm_response}
            namespace = {'openpyxl': openpyxl, 'pd': pd, 'BytesIO': BytesIO}
            exec(code, namespace)
            transform_fn = namespace.get('transform')
            if not transform_fn:
                raise RuntimeError('生成的代码中没有 transform 函数')
            result_df = transform_fn(file_content, sheet_name)
            if not isinstance(result_df, pd.DataFrame):
                raise TypeError(f'transform 返回类型不正确: {type(result_df)}')
            if len(result_df) == 0:
                raise ValueError('transform 返回空 DataFrame')
            n_cols = len(result_df.columns)
            n_rows = len(result_df)
            null_ratio = result_df.isnull().sum().sum() / (n_cols * n_rows) if n_cols * n_rows > 0 else 0
            if null_ratio > 0.5:
                raise ValueError(f'结果中空值占比过高 ({null_ratio:.1%})，共 {n_rows}行×{n_cols}列，可能是不同表块被错误合并了。请确保每个独立表块按各自的列范围单独提取，再按处理名+重复编号合并。')
            print(f'[LLM Tidy] 成功！{n_rows}行 × {n_cols}列, 空值率 {null_ratio:.1%}')
            return {'success': True, 'description': description, 'tables_found': tables_found, 'result_df': result_df, 'error': None, 'llm_response': llm_response}
        except json.JSONDecodeError as e:
            print(f'[LLM Tidy] JSON 解析失败: {e}')
            if attempt < max_retries:
                prompt = f'你上次返回的不是合法 JSON，解析错误: {str(e)}\n你的原始回复前100字符: {last_response[:100]}\n请严格遵循格式要求，只返回 JSON 对象，不要加 ```json 标记或其他文字。'
                continue
            return {'success': False, 'description': description, 'tables_found': tables_found, 'result_df': None, 'error': f'JSON 解析失败: {str(e)}', 'llm_response': last_response}
        except Exception as e:
            error_msg = str(e)
            tb = traceback.format_exc()
            print(f'[LLM Tidy] 第 {attempt + 1} 次失败: {error_msg}')
            if attempt < max_retries:
                prompt = f"""你上次生成的代码执行出错了，请根据以下错误信息修正。\n\n错误信息: {error_msg}\n\n错误追踪 (最后500字符):\n{tb[-500:]}\n\n你上次生成的代码:\n```python\n{(code[:3000] if code else '(无)')}\n```\n\n关键提示:\n- 如果是"空值占比过高"的错误，说明不同列区域的表块被混在了同一行，请确保每个表块按自己的列范围独立提取\n- 如果是 KeyError 或 IndexError，检查行列索引是否正确\n- 确保填充合并单元格后再读取数据\n\n请返回修正后的完整 JSON 对象（包含 description、tables_found、code）。"""
                continue
            return {'success': False, 'description': description, 'tables_found': tables_found, 'result_df': None, 'error': f'执行失败 (已重试 {max_retries} 次): {error_msg}', 'llm_response': last_response}

# ==============================
# Flask ??
# ==============================

app = Flask(__name__)
PORT = 5678
INDEX_HTML = """<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Excel 智能整理工具</title>
  <script src="https://cdn.tailwindcss.com"></script>
  <link rel="preconnect" href="https://fonts.googleapis.com" />
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin />
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=Noto+Sans+SC:wght@400;500;700&display=swap" rel="stylesheet" />
  <link href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:opsz,wght,FILL,GRAD@24,500,0,0" rel="stylesheet" />
  <style>
    :root { --grad-start:#4f7cff; --grad-end:#6a5cff; --ink:#12203a; }
    html,body{margin:0;padding:0;font-family:"Inter","Noto Sans SC",sans-serif;color:var(--ink);background:radial-gradient(circle at 8% 12%,#e8eeff 0%,#f8f9ff 42%,#eef3ff 100%);height:100vh;overflow:hidden;}
    .material-symbols-outlined{font-variation-settings:"FILL" 0,"wght" 500,"GRAD" 0,"opsz" 24;}
    .glass{background:rgba(255,255,255,.8);backdrop-filter:blur(8px);}
    .drop-zone.dragover{border-color:#4f7cff;background:#eef3ff;transform:translateY(-1px);}
    /* Full-height flex layout */
    .app-shell{display:flex;flex-direction:column;height:100vh;padding:16px;box-sizing:border-box;gap:12px;}
    .app-shell>.app-header{flex-shrink:0;}
    .app-shell>.app-body{flex:1;min-height:0;display:grid;grid-template-columns:320px 1fr;gap:16px;}
    .app-body>aside{overflow-y:auto;}
    .app-body>section{display:flex;flex-direction:column;min-height:0;overflow:hidden;}
    .data-host{display:flex;flex-direction:column;flex:1;min-height:0;}
    .table-wrap{overflow:auto;border:1px solid #e8edff;border-radius:14px;background:#fff;flex:1;min-height:0;}
    table{border-collapse:collapse;width:100%;min-width:640px;}
    th,td{border-bottom:1px solid #eef2ff;border-right:1px solid #f2f5ff;padding:9px 12px;font-size:13px;white-space:nowrap;text-align:left;}
    th{position:sticky;top:0;z-index:1;background:#f4f7ff;color:#263b7f;font-weight:700;}
    .editable-cell{cursor:text;transition:all .15s ease;position:relative;}
    .editable-cell:hover{background:#f5f7ff;}
    .cell-editing{outline:2px solid #4f7cff;background:#f0f4ff !important;}
    .cell-save-success{animation:cell-save-flash .6s ease;}
    @keyframes cell-save-flash{0%{box-shadow:inset 0 0 0 0 #22c55e;}40%{box-shadow:inset 0 0 0 2px #22c55e;}100%{box-shadow:inset 0 0 0 0 #22c55e;}}
    .edit-hint{display:inline-flex;align-items:center;gap:4px;font-size:11px;color:#94a3b8;}
    /* View tabs */
    .view-tabs{display:flex;gap:4px;background:#f1f3f9;border-radius:12px;padding:4px;margin-bottom:8px;flex-shrink:0;}
    .view-tab{flex:1;display:inline-flex;align-items:center;justify-content:center;gap:6px;padding:8px 12px;border:none;border-radius:10px;font-size:14px;font-weight:600;cursor:pointer;transition:all .2s ease;color:#6b7280;background:transparent;}
    .view-tab:hover{color:#4f7cff;background:rgba(79,124,255,.06);}
    .view-tab.active{background:#fff;color:#4f7cff;box-shadow:0 1px 4px rgba(0,0,0,.08);}
    .view-tab .material-symbols-outlined{font-size:18px;}
    .view-tab .badge{display:inline-flex;align-items:center;min-width:20px;height:20px;border-radius:10px;font-size:10px;font-weight:700;padding:0 6px;background:#e0e7ff;color:#4f46e5;margin-left:2px;}
    .view-tab.active .badge{background:#4f7cff;color:#fff;}
    .view-panel{display:none;flex-direction:column;flex:1;min-height:0;}
    .view-panel.active{display:flex;}
    .panel-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:8px;flex-shrink:0;gap:8px;}
    .toast{animation:toast-in .18s ease-out;} @keyframes toast-in{from{opacity:0;transform:translateY(8px)}to{opacity:1;transform:translateY(0)}}
    @media(max-width:1024px){.app-body{grid-template-columns:1fr !important;}}
  </style>
</head>
<body>
  <div class="app-shell">
    <header class="app-header rounded-2xl p-5 text-white shadow-xl" style="background:linear-gradient(135deg,var(--grad-start),var(--grad-end));">
      <div class="flex items-center gap-3">
        <span class="material-symbols-outlined text-3xl">table_chart</span>
        <div>
          <h1 class="text-2xl font-extrabold tracking-tight">Excel 智能整理工具</h1>
          <p class="mt-0.5 text-sm text-blue-100">上传 → 选择工作表 → 一键整理 → 导出</p>
        </div>
      </div>
    </header>
    <main class="app-body">
      <aside class="rounded-2xl border border-white/60 glass p-4 shadow-lg">
          <div class="mb-4"><h2 class="font-bold text-base">控制面板</h2><p class="text-sm text-slate-500 mt-1">支持 .xlsx 和 .xls 格式</p></div>
          <div id="dropZone" class="drop-zone mb-4 rounded-2xl border-2 border-dashed border-indigo-200 bg-indigo-50/60 p-5 transition-all cursor-pointer">
            <input id="fileInput" type="file" class="hidden" accept=".xls,.xlsx" />
            <div class="flex flex-col items-center text-center gap-2">
              <span class="material-symbols-outlined text-[34px] text-indigo-600">upload_file</span>
              <div class="font-semibold text-indigo-700">拖拽 Excel 文件到此处</div><div class="text-xs text-slate-500">或点击选择文件</div>
              <div id="fileName" class="text-xs text-slate-400"></div>
            </div>
          </div>
          <button id="tidyBtn" class="w-full rounded-xl py-3.5 px-4 text-white font-semibold shadow-lg disabled:opacity-40 disabled:cursor-not-allowed" style="background:linear-gradient(135deg,var(--grad-start),var(--grad-end));" disabled>
            <span class="inline-flex items-center gap-2"><span class="material-symbols-outlined text-[20px]">auto_fix_high</span>一键智能整理</span>
          </button>
          <details class="mt-4 rounded-xl border border-indigo-100 bg-white/90 p-3">
            <summary class="cursor-pointer select-none text-sm font-semibold text-indigo-700">LLM 兜底配置</summary>
            <div class="mt-3 space-y-3">
              <div><label class="block text-xs text-slate-500 mb-1">API 密钥</label><input id="apiKey" type="password" placeholder="仅本地整理失败时使用" class="w-full rounded-lg border border-slate-200 px-3 py-2 text-sm focus:outline-none focus:ring-2 focus:ring-indigo-300" /></div>
              <div><label class="block text-xs text-slate-500 mb-1">模型</label><input id="model" type="text" value="deepseek-ai/DeepSeek-V3" class="w-full rounded-lg border border-slate-200 px-3 py-2 text-sm focus:outline-none focus:ring-2 focus:ring-indigo-300" /></div>
              <div><label class="block text-xs text-slate-500 mb-1">API 地址</label><input id="apiBase" type="text" value="https://api.siliconflow.cn/v1" class="w-full rounded-lg border border-slate-200 px-3 py-2 text-sm focus:outline-none focus:ring-2 focus:ring-indigo-300" /></div>
            </div>
            <div class="mt-3 text-right"><button id="saveSettingsBtn" class="inline-flex items-center rounded-md bg-blue-600 px-3 py-1.5 text-xs font-semibold text-white hover:bg-blue-700">保存设置</button></div>
          </details>
          <div class="mt-4 rounded-xl border border-indigo-100 bg-white/90 p-3"><div class="text-xs text-slate-500">状态</div><div id="statusText" class="mt-1 text-sm font-medium text-slate-700">等待上传文件</div><div id="statusMeta" class="text-xs text-slate-500 mt-1"></div></div>
        </aside>
        <section class="rounded-2xl border border-white/70 bg-white/90 p-4 shadow-lg">
          <div class="data-host">
            <div class="view-tabs">
              <button id="tabRaw" class="view-tab active" onclick="switchTab('raw')"><span class="material-symbols-outlined">table_view</span>原始数据<span id="rawBadge" class="badge" style="display:none;"></span></button>
              <button id="tabTidy" class="view-tab" onclick="switchTab('tidy')"><span class="material-symbols-outlined">auto_fix_high</span>整理结果<span id="tidyBadge" class="badge" style="display:none;"></span></button>
            </div>
            <div id="panelRaw" class="view-panel active">
              <div class="panel-header"><h3 class="font-semibold text-slate-700">原始数据预览</h3><div class="flex items-center gap-3"><span class="edit-hint"><span class="material-symbols-outlined" style="font-size:14px">edit_note</span>双击单元格可编辑</span><span id="rawMeta" class="text-xs text-slate-500"></span></div></div>
              <div id="rawTable" class="table-wrap"></div>
            </div>
            <div id="panelTidy" class="view-panel">
              <div class="panel-header"><h3 class="font-semibold text-slate-700">整理后数据预览</h3><div class="flex items-center gap-2"><span class="edit-hint" id="tidyEditHint" style="display:none;"><span class="material-symbols-outlined" style="font-size:14px">edit_note</span>双击可编辑</span><span id="tidyMeta" class="text-xs text-slate-500"></span><button id="exportBtn" class="rounded-lg bg-emerald-500 px-3 py-2 text-sm text-white font-semibold disabled:opacity-40 disabled:cursor-not-allowed" disabled><span class="inline-flex items-center gap-1"><span class="material-symbols-outlined text-[18px]">download</span>导出 Excel</span></button></div></div>
              <div id="tidyTable" class="table-wrap"></div>
            </div>
          </div>
        </section>
      </main>
    </div>
  <div id="sheetModal" class="fixed inset-0 z-40 hidden items-center justify-center bg-slate-900/45 p-4"><div class="w-full max-w-md rounded-2xl bg-white p-5 shadow-2xl"><div class="flex items-center gap-2"><span class="material-symbols-outlined text-indigo-600">view_array</span><h3 class="text-lg font-bold">选择工作表</h3></div><p class="mt-2 text-sm text-slate-500">检测到多个工作表，请选择一个继续。</p><div id="sheetList" class="mt-4 grid gap-2"></div><div class="mt-4 text-right"><button id="sheetCancel" class="rounded-lg border border-slate-200 px-3 py-2 text-sm">关闭</button></div></div></div>
  <div id="toastStack" class="fixed right-4 top-4 z-50 space-y-2"></div>
  <script>
    const state = { dataId: null, selectedSheet: null, sheets: [], rawColumns: [], rawRows: [], tidyColumns: [], tidyRows: [] };
    const el = {
      dropZone: document.getElementById('dropZone'),
      fileInput: document.getElementById('fileInput'),
      fileName: document.getElementById('fileName'),
      tidyBtn: document.getElementById('tidyBtn'),
      exportBtn: document.getElementById('exportBtn'),
      rawTable: document.getElementById('rawTable'),
      tidyTable: document.getElementById('tidyTable'),
      rawMeta: document.getElementById('rawMeta'),
      tidyMeta: document.getElementById('tidyMeta'),
      statusText: document.getElementById('statusText'),
      statusMeta: document.getElementById('statusMeta'),
      sheetModal: document.getElementById('sheetModal'),
      sheetList: document.getElementById('sheetList'),
      sheetCancel: document.getElementById('sheetCancel'),
      apiKey: document.getElementById('apiKey'),
      model: document.getElementById('model'),
      apiBase: document.getElementById('apiBase'),
      saveSettingsBtn: document.getElementById('saveSettingsBtn'),
      toastStack: document.getElementById('toastStack'),
      tidyEditHint: document.getElementById('tidyEditHint'),
      tabRaw: document.getElementById('tabRaw'),
      tabTidy: document.getElementById('tabTidy'),
      panelRaw: document.getElementById('panelRaw'),
      panelTidy: document.getElementById('panelTidy'),
      rawBadge: document.getElementById('rawBadge'),
      tidyBadge: document.getElementById('tidyBadge')
    };
    const setStatus = (text, meta = '') => { el.statusText.textContent = text; el.statusMeta.textContent = meta; };
    const toast = (message, type = 'info') => {
      const color = { success: 'bg-emerald-500', error: 'bg-rose-500', info: 'bg-slate-800' }[type] || 'bg-slate-800';
      const node = document.createElement('div');
      node.className = `toast ${color} text-white text-sm px-4 py-2 rounded-lg shadow-lg`;
      node.textContent = message;
      el.toastStack.appendChild(node);
      setTimeout(() => node.remove(), 2600);
    };
    const esc = (value) => String(value).replaceAll('&', '&amp;').replaceAll('<', '&lt;').replaceAll('>', '&gt;').replaceAll('"', '&quot;').replaceAll("'", '&#39;');

    function switchTab(tab) {
      const isRaw = tab === 'raw';
      el.tabRaw.classList.toggle('active', isRaw);
      el.tabTidy.classList.toggle('active', !isRaw);
      el.panelRaw.classList.toggle('active', isRaw);
      el.panelTidy.classList.toggle('active', !isRaw);
    }

    function updateBadges() {
      if (state.rawRows.length > 0) { el.rawBadge.textContent = state.rawRows.length; el.rawBadge.style.display = 'inline-flex'; } else { el.rawBadge.style.display = 'none'; }
      if (state.tidyRows.length > 0) { el.tidyBadge.textContent = state.tidyRows.length; el.tidyBadge.style.display = 'inline-flex'; } else { el.tidyBadge.style.display = 'none'; }
    }

    function renderTable(container, columns, rows, emptyText, options = {}) {
      const editable = Boolean(options.editable);
      if (!columns || !columns.length) {
        container.innerHTML = `<div class="p-6 text-sm text-slate-400 text-center">${emptyText}</div>`;
        return;
      }
      const head = columns.map((c) => `<th>${esc(c)}</th>`).join('');
      const body = rows.map((row, rowIndex) => {
        const cells = columns.map((col, colIndex) => {
          const value = esc(row?.[col] ?? '');
          if (!editable) {
            return `<td>${value}</td>`;
          }
          return `<td class="editable-cell" data-row-index="${rowIndex}" data-col-index="${colIndex}" title="双击编辑">${value}</td>`;
        }).join('');
        return `<tr>${cells}</tr>`;
      }).join('');
      container.innerHTML = `<table><thead><tr>${head}</tr></thead><tbody>${body || `<tr><td colspan="${columns.length}" class="text-slate-400 text-center">暂无数据</td></tr>`}</tbody></table>`;
      if (editable) {
        bindEditableTable(container);
      }
    }

    function cleanupEditingCell(cell) {
      cell.removeAttribute('contenteditable');
      cell.classList.remove('cell-editing');
      delete cell.dataset.editing;
      delete cell.dataset.saving;
      delete cell.dataset.originalValue;
    }

    function startCellEdit(cell) {
      if (!state.dataId || cell.dataset.saving === '1') return;
      if (cell.dataset.editing === '1') return;
      cell.dataset.originalValue = cell.textContent ?? '';
      cell.dataset.editing = '1';
      cell.setAttribute('contenteditable', 'true');
      cell.classList.add('cell-editing');
      cell.focus();
      const selection = window.getSelection();
      const range = document.createRange();
      range.selectNodeContents(cell);
      range.collapse(false);
      selection.removeAllRanges();
      selection.addRange(range);
    }

    async function finishCellEdit(cell) {
      if (cell.dataset.editing !== '1' || cell.dataset.saving === '1') return;
      const originalValue = cell.dataset.originalValue ?? '';
      const currentValue = cell.textContent ?? '';
      const rowIndex = Number(cell.dataset.rowIndex);
      const colIndex = Number(cell.dataset.colIndex);
      const columnName = state.tidyColumns[colIndex];
      const apiUrl = '/api/update_cell';
      if (!Number.isInteger(rowIndex) || !Number.isInteger(colIndex) || !columnName) {
        cleanupEditingCell(cell);
        return;
      }
      if (currentValue === originalValue) {
        cleanupEditingCell(cell);
        return;
      }
      cell.dataset.saving = '1';
      try {
        const data = await fetchJson(apiUrl, {
          method: 'POST',
          headers: { 'Content-Type': 'application/json' },
          body: JSON.stringify({ data_id: state.dataId, row_index: rowIndex, column_name: columnName, new_value: currentValue })
        });
        const savedValue = data.updated_value == null ? '' : String(data.updated_value);
        cell.textContent = savedValue;
        if (!state.tidyRows[rowIndex]) state.tidyRows[rowIndex] = {};
        state.tidyRows[rowIndex][columnName] = savedValue;
        cell.classList.remove('cell-save-success');
        void cell.offsetWidth;
        cell.classList.add('cell-save-success');
        setTimeout(() => cell.classList.remove('cell-save-success'), 700);
      } catch (err) {
        cell.textContent = originalValue;
        toast(err.message || '单元格保存失败', 'error');
      } finally {
        cleanupEditingCell(cell);
      }
    }

    function cancelCellEdit(cell) {
      if (cell.dataset.editing !== '1') return;
      cell.textContent = cell.dataset.originalValue ?? '';
      cleanupEditingCell(cell);
    }

    function bindEditableTable(container) {
      if (container.dataset.editableBound === '1') return;
      container.dataset.editableBound = '1';
      container.addEventListener('dblclick', (event) => {
        const cell = event.target.closest('td.editable-cell');
        if (!cell) return;
        startCellEdit(cell);
      });
      container.addEventListener('keydown', (event) => {
        const cell = event.target.closest('td.editable-cell');
        if (!cell || cell.dataset.editing !== '1') return;
        if (event.key === 'Enter') {
          event.preventDefault();
          cell.blur();
          return;
        }
        if (event.key === 'Escape') {
          event.preventDefault();
          cancelCellEdit(cell);
          return;
        }
        if (event.key === 'Tab') {
          event.preventDefault();
          cell.blur();
          const allCells = Array.from(container.querySelectorAll('td.editable-cell'));
          const currentIdx = allCells.indexOf(cell);
          if (currentIdx < 0) return;
          const nextIdx = event.shiftKey ? currentIdx - 1 : currentIdx + 1;
          if (nextIdx >= 0 && nextIdx < allCells.length) {
            setTimeout(() => startCellEdit(allCells[nextIdx]), 50);
          }
        }
      });
      container.addEventListener('blur', (event) => {
        const cell = event.target.closest('td.editable-cell');
        if (!cell) return;
        finishCellEdit(cell);
      }, true);
    }

    async function fetchJson(url, opt = {}) {
      const res = await fetch(url, opt);
      const data = await res.json().catch(() => ({}));
      if (!res.ok) throw new Error(data.error || `请求失败: ${res.status}`);
      return data;
    }

    async function loadSettings() {
      try {
        const data = await fetchJson('/api/settings');
        el.apiKey.value = data.api_key || '';
        el.model.value = data.model || '';
        el.apiBase.value = data.api_base || '';
      } catch (err) {
        setStatus('设置读取失败', err.message || '请手动填写 LLM 配置');
      }
    }

    async function saveSettings() {
      try {
        const payload = {
          api_key: (el.apiKey.value || '').trim(),
          model: (el.model.value || '').trim(),
          api_base: (el.apiBase.value || '').trim()
        };
        await fetchJson('/api/settings', {
          method: 'POST',
          headers: { 'Content-Type': 'application/json' },
          body: JSON.stringify(payload)
        });
        toast('设置已保存', 'success');
        setStatus('设置已保存', '已写入 settings.json');
      } catch (err) {
        toast(err.message || '设置保存失败', 'error');
      }
    }

    async function uploadFile(file) {
      const fd = new FormData();
      fd.append('file', file);
      setStatus('正在上传并解析...');
      const data = await fetchJson('/api/upload', { method: 'POST', body: fd });
      state.dataId = data.data_id;
      state.selectedSheet = data.selected_sheet || null;
      state.sheets = data.sheets || [];
      state.rawColumns = data.columns || [];
      state.rawRows = data.preview || [];
      state.tidyColumns = [];
      state.tidyRows = [];
      renderTable(el.rawTable, state.rawColumns, state.rawRows, '上传后将在此显示原始数据预览');
      renderTable(el.tidyTable, [], [], '点击”一键智能整理”查看结果预览', { editable: true });
      el.rawMeta.textContent = `${data.rows || 0} 行 / ${state.rawColumns.length} 列`;
      el.tidyMeta.textContent = '';
      el.exportBtn.disabled = true;
      el.tidyBtn.disabled = false;
      el.tidyEditHint.style.display = 'none';
      updateBadges();
      switchTab('raw');
      setStatus('上传完成', `${data.filename || ''} | 工作表: ${state.selectedSheet || '-'}`);
      toast('上传完成', 'success');
      if (state.sheets.length > 1) showSheetModal(state.sheets);
    }

    function showSheetModal(sheets) {
      el.sheetList.innerHTML = '';
      sheets.forEach((name) => {
        const btn = document.createElement('button');
        btn.className = 'rounded-lg border border-indigo-100 px-3 py-2 text-left text-sm hover:bg-indigo-50';
        btn.textContent = name;
        btn.onclick = async () => {
          await loadSheet(name);
          hideSheetModal();
        };
        el.sheetList.appendChild(btn);
      });
      el.sheetModal.classList.remove('hidden');
      el.sheetModal.classList.add('flex');
    }

    function hideSheetModal() {
      el.sheetModal.classList.add('hidden');
      el.sheetModal.classList.remove('flex');
    }

    async function loadSheet(sheetName) {
      if (!state.dataId) return;
      setStatus('正在加载工作表...', sheetName);
      const data = await fetchJson('/api/load_sheet', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ data_id: state.dataId, sheet_name: sheetName })
      });
      state.selectedSheet = data.selected_sheet;
      state.rawColumns = data.columns || [];
      state.rawRows = data.preview || [];
      state.tidyColumns = [];
      state.tidyRows = [];
      renderTable(el.rawTable, state.rawColumns, state.rawRows, '上传后将在此显示原始数据预览');
      renderTable(el.tidyTable, [], [], '点击”一键智能整理”查看结果预览', { editable: true });
      el.rawMeta.textContent = `${data.rows || 0} 行 / ${state.rawColumns.length} 列`;
      el.tidyMeta.textContent = '';
      el.exportBtn.disabled = true;
      el.tidyEditHint.style.display = 'none';
      updateBadges();
      switchTab('raw');
      setStatus('工作表切换完成', `当前: ${state.selectedSheet || '-'}`);
      toast(`已切换工作表: ${sheetName}`, 'info');
    }

    async function tidyData() {
      if (!state.dataId) {
        toast('请先上传 Excel 文件', 'error');
        return;
      }
      el.tidyBtn.disabled = true;
      setStatus('正在执行双引擎整理...');
      try {
        const data = await fetchJson('/api/tidy', {
          method: 'POST',
          headers: { 'Content-Type': 'application/json' },
          body: JSON.stringify({
            data_id: state.dataId,
            sheet_name: state.selectedSheet,
            api_key: (el.apiKey.value || '').trim(),
            model: (el.model.value || '').trim(),
            api_base: (el.apiBase.value || '').trim(),
            drop_avg: true
          })
        });
        state.tidyColumns = data.columns || [];
        state.tidyRows = data.preview || [];
        renderTable(el.tidyTable, state.tidyColumns, state.tidyRows, '整理后数据预览', { editable: true });
        el.tidyMeta.textContent = `${data.total_rows || 0} 行 / ${state.tidyColumns.length} 列`;
        el.exportBtn.disabled = state.tidyColumns.length === 0;
        el.tidyEditHint.style.display = state.tidyColumns.length > 0 ? 'inline-flex' : 'none';
        updateBadges();
        switchTab('tidy');
        const label = data.engine === 'smart_tidy' ? '本地 SmartTidy 引擎' : 'LLM 兜底引擎';
        setStatus('整理完成', `引擎: ${label}`);
        toast(`整理完成（${label}）`, 'success');
      } catch (err) {
        setStatus('整理失败', err.message || '未知错误');
        toast(err.message || '整理失败', 'error');
      } finally {
        el.tidyBtn.disabled = false;
      }
    }

    async function exportResult() {
      if (!state.dataId) {
        toast('没有可导出的结果', 'error');
        return;
      }
      try {
        const res = await fetch('/api/export', {
          method: 'POST',
          headers: { 'Content-Type': 'application/json' },
          body: JSON.stringify({ data_id: state.dataId })
        });
        if (!res.ok) {
          const data = await res.json().catch(() => ({}));
          throw new Error(data.error || '导出失败');
        }
        const blob = await res.blob();
        const url = window.URL.createObjectURL(blob);
        const link = document.createElement('a');
        link.href = url;
        const dispo = res.headers.get('Content-Disposition') || '';
        const match = dispo.match(/filename="?([^";]+)"?/i);
        link.download = match ? decodeURIComponent(match[1]) : 'tidy_result.xlsx';
        document.body.appendChild(link);
        link.click();
        link.remove();
        window.URL.revokeObjectURL(url);
        toast('导出完成', 'success');
      } catch (err) {
        toast(err.message || '导出失败', 'error');
      }
    }

    async function addRawRow() {
      if (!state.dataId) { toast('请先上传文件', 'error'); return; }
      try {
        const data = await fetchJson('/api/add_raw_row', {
          method: 'POST',
          headers: { 'Content-Type': 'application/json' },
          body: JSON.stringify({ data_id: state.dataId })
        });
        state.rawColumns = data.columns || state.rawColumns;
        state.rawRows = data.preview || [];
        renderTable(el.rawTable, state.rawColumns, state.rawRows, '上传后将在此显示原始数据预览');
        el.rawMeta.textContent = `${data.total_rows || state.rawRows.length} 行 / ${state.rawColumns.length} 列`;
        toast('已添加新行', 'success');
      } catch (err) {
        toast(err.message || '添加行失败', 'error');
      }
    }

    async function deleteRawRow(rowIndex) {
      if (!state.dataId) return;
      try {
        const data = await fetchJson('/api/delete_raw_row', {
          method: 'POST',
          headers: { 'Content-Type': 'application/json' },
          body: JSON.stringify({ data_id: state.dataId, row_index: rowIndex })
        });
        state.rawColumns = data.columns || state.rawColumns;
        state.rawRows = data.preview || [];
        renderTable(el.rawTable, state.rawColumns, state.rawRows, '上传后将在此显示原始数据预览');
        el.rawMeta.textContent = `${data.total_rows || state.rawRows.length} 行 / ${state.rawColumns.length} 列`;
        toast('已删除行', 'info');
      } catch (err) {
        toast(err.message || '删除行失败', 'error');
      }
    }

    function bindEvents() {
      el.dropZone.addEventListener('click', () => el.fileInput.click());
      el.fileInput.addEventListener('change', async (event) => {
        const file = event.target.files?.[0];
        if (!file) return;
        el.fileName.textContent = file.name;
        try {
          await uploadFile(file);
        } catch (err) {
          setStatus('上传失败', err.message || '未知错误');
          toast(err.message || '上传失败', 'error');
        }
      });
      ['dragenter', 'dragover'].forEach((name) => el.dropZone.addEventListener(name, (ev) => {
        ev.preventDefault();
        ev.stopPropagation();
        el.dropZone.classList.add('dragover');
      }));
      ['dragleave', 'drop'].forEach((name) => el.dropZone.addEventListener(name, (ev) => {
        ev.preventDefault();
        ev.stopPropagation();
        el.dropZone.classList.remove('dragover');
      }));
      el.dropZone.addEventListener('drop', async (event) => {
        const file = event.dataTransfer?.files?.[0];
        if (!file) return;
        el.fileName.textContent = file.name;
        try {
          await uploadFile(file);
        } catch (err) {
          setStatus('上传失败', err.message || '未知错误');
          toast(err.message || '上传失败', 'error');
        }
      });
      el.tidyBtn.addEventListener('click', tidyData);
      el.exportBtn.addEventListener('click', exportResult);
      el.saveSettingsBtn.addEventListener('click', saveSettings);
      el.sheetCancel.addEventListener('click', hideSheetModal);
    }

    async function init() {
      bindEvents();
      renderTable(el.rawTable, [], [], '上传后将在此显示原始数据预览');
      renderTable(el.tidyTable, [], [], '点击”一键智能整理”查看结果预览', { editable: true });
      await loadSettings();
    }

    init();
  </script>
</body>
</html>"""


@app.get('/')
def index() -> Response:
    return Response(INDEX_HTML, mimetype='text/html')


@app.get('/api/settings')
def api_get_settings():
    return jsonify({key: app_settings.get(key, '') for key in ('api_key', 'model', 'api_base')})


@app.post('/api/settings')
def api_save_settings():
    payload = request.json or {}
    if not isinstance(payload, dict):
        return jsonify({'error': '请求体格式错误'}), 400
    settings = {
        'api_key': (payload.get('api_key') or '').strip(),
        'model': (payload.get('model') or '').strip() or DEFAULT_MODEL,
        'api_base': (payload.get('api_base') or '').strip() or DEFAULT_API_BASE,
    }
    try:
        _save_settings(settings)
    except Exception as exc:
        traceback.print_exc()
        return jsonify({'error': f'保存设置失败: {str(exc)}'}), 500
    app_settings.update(settings)
    return jsonify({'success': True, **settings})


@app.post('/api/upload')
def api_upload():
    if 'file' not in request.files:
        return jsonify({'error': '未找到上传文件'}), 400
    file = request.files['file']
    filename = file.filename or ''
    if not filename:
        return jsonify({'error': '文件名为空'}), 400
    ext = filename.lower().rsplit('.', 1)
    ext = f'.{ext[-1]}' if len(ext) == 2 else ''
    if ext not in {'.xlsx', '.xls'}:
        return jsonify({'error': '仅支持 .xls / .xlsx 格式'}), 400
    try:
        file_content = file.read()
        if not file_content:
            return jsonify({'error': '文件内容为空'}), 400
        sheet_names = _list_sheet_names(file_content, ext)
        selected_sheet = sheet_names[0] if sheet_names else None
        raw_df = _read_sheet_dataframe(file_content, ext, selected_sheet)
        xlsx_content = file_content if ext == '.xlsx' else None
        if ext == '.xls':
            try:
                xlsx_content = _convert_xls_to_xlsx_bytes(file_content)
            except Exception:
                xlsx_content = None
        data_id = str(uuid.uuid4())
        raw_file_store[data_id] = {'filename': filename, 'ext': ext, 'content': file_content, 'xlsx_content': xlsx_content, 'sheet_names': sheet_names, 'selected_sheet': selected_sheet}
        data_store[data_id] = {'raw_df': raw_df, 'tidy_df': None, 'sheet_name': selected_sheet}
        return jsonify({'data_id': data_id, 'filename': filename, 'sheets': sheet_names, 'selected_sheet': selected_sheet, 'columns': [str(c) for c in raw_df.columns], 'rows': int(len(raw_df)), 'preview': dataframe_preview(raw_df)})
    except Exception as exc:
        traceback.print_exc()
        return jsonify({'error': f'上传解析失败: {str(exc)}'}), 500


@app.post('/api/load_sheet')
def api_load_sheet():
    payload = request.json or {}
    data_id = payload.get('data_id')
    sheet_name = payload.get('sheet_name')
    if not data_id or data_id not in raw_file_store:
        return jsonify({'error': '数据已过期，请重新上传'}), 400
    if not sheet_name:
        return jsonify({'error': '缺少 sheet_name'}), 400
    file_info = raw_file_store[data_id]
    if sheet_name not in file_info.get('sheet_names', []):
        return jsonify({'error': '工作表不存在'}), 400
    try:
        raw_df = _read_sheet_dataframe(file_info['content'], file_info['ext'], sheet_name)
        data_store[data_id] = {'raw_df': raw_df, 'tidy_df': None, 'sheet_name': sheet_name}
        file_info['selected_sheet'] = sheet_name
        return jsonify({'data_id': data_id, 'selected_sheet': sheet_name, 'columns': [str(c) for c in raw_df.columns], 'rows': int(len(raw_df)), 'preview': dataframe_preview(raw_df)})
    except Exception as exc:
        traceback.print_exc()
        return jsonify({'error': f'切换工作表失败: {str(exc)}'}), 500


@app.post('/api/tidy')
def api_tidy():
    payload = request.json or {}
    data_id = payload.get('data_id')
    sheet_name = payload.get('sheet_name')
    drop_avg = bool(payload.get('drop_avg', True))
    api_key = (payload.get('api_key') or '').strip()
    model = (payload.get('model') or DEFAULT_MODEL).strip() or DEFAULT_MODEL
    api_base = (payload.get('api_base') or DEFAULT_API_BASE).strip() or DEFAULT_API_BASE
    if not data_id or data_id not in raw_file_store:
        return jsonify({'error': '数据已过期，请重新上传'}), 400
    file_info = raw_file_store[data_id]
    sheet_name = sheet_name or file_info.get('selected_sheet')
    if sheet_name and sheet_name not in file_info.get('sheet_names', []):
        return jsonify({'error': '工作表不存在'}), 400
    xlsx_content = file_info.get('xlsx_content')
    if not xlsx_content:
        try:
            if file_info.get('ext') == '.xls':
                xlsx_content = _convert_xls_to_xlsx_bytes(file_info['content'])
                file_info['xlsx_content'] = xlsx_content
            else:
                xlsx_content = file_info['content']
        except Exception as exc:
            xlsx_content = None
            smart_error = f'格式转换失败: {str(exc)}'
        else:
            smart_error = None
    else:
        smart_error = None

    engine = None
    tidy_df = None
    detail_message = ''
    if xlsx_content:
        try:
            scan_result = scan_excel_structure(xlsx_content, sheet_name)
            if isinstance(scan_result, dict) and 'error' not in scan_result:
                tidy_df = execute_smart_tidy(scan_result, {'drop_avg': drop_avg, 'output_format': 'wide'})
                if isinstance(tidy_df, pd.DataFrame) and not tidy_df.empty:
                    engine = 'smart_tidy'
                    detail_message = '已使用本地 SmartTidy 引擎完成整理'
                else:
                    smart_error = '本地 SmartTidy 结果为空'
            else:
                smart_error = (scan_result or {}).get('error', '本地 SmartTidy 扫描失败')
        except Exception as exc:
            traceback.print_exc()
            smart_error = f'本地 SmartTidy 异常: {str(exc)}'

    if tidy_df is None or tidy_df.empty:
        if not api_key:
            return jsonify({'error': f"{smart_error or '本地解析失败'}，且未提供 API Key，无法启用 LLM 兜底"}), 400
        if not xlsx_content:
            return jsonify({'error': f"无法调用 LLM 兜底：{smart_error or '本地解析失败'}"}), 400
        llm_result = analyze_and_transform(file_content=xlsx_content, api_key=api_key, sheet_name=sheet_name, api_base=api_base, model=model, max_retries=3)
        if not llm_result.get('success'):
            err = llm_result.get('error') or 'LLM 执行失败'
            if smart_error:
                err = f'{smart_error}；LLM: {err}'
            return jsonify({'error': err}), 400
        tidy_df = llm_result.get('result_df')
        if not isinstance(tidy_df, pd.DataFrame) or tidy_df.empty:
            return jsonify({'error': 'LLM 返回空结果'}), 400
        engine = 'llm_tidy'
        detail_message = '本地解析失败后，已由 LLM 兜底完成整理'

    tidy_df = sanitize_dataframe(tidy_df)
    if data_id not in data_store:
        data_store[data_id] = {}
    data_store[data_id]['tidy_df'] = tidy_df
    data_store[data_id]['sheet_name'] = sheet_name
    return jsonify({'success': True, 'engine': engine, 'message': detail_message, 'columns': [str(c) for c in tidy_df.columns], 'total_rows': int(len(tidy_df)), 'preview': dataframe_preview(tidy_df), 'sheet_name': sheet_name})


@app.post('/api/update_raw_cell')
def api_update_raw_cell():
    """编辑原始数据的单元格"""
    payload = request.json or {}
    data_id = payload.get('data_id')
    if not data_id or data_id not in data_store:
        return jsonify({'error': '数据不存在或已过期'}), 400
    raw_df = data_store[data_id].get('raw_df')
    if raw_df is None or not isinstance(raw_df, pd.DataFrame) or raw_df.empty:
        return jsonify({'error': '没有可编辑的原始数据'}), 400
    try:
        row_index = int(payload.get('row_index'))
    except (TypeError, ValueError):
        return jsonify({'error': 'row_index 必须是整数'}), 400
    column_name = str(payload.get('column_name') or '').strip()
    if not column_name:
        return jsonify({'error': 'column_name 不能为空'}), 400
    if column_name not in raw_df.columns:
        return jsonify({'error': f'列不存在: {column_name}'}), 400
    if row_index < 0 or row_index >= len(raw_df):
        return jsonify({'error': f'row_index 越界: {row_index}'}), 400
    new_value = payload.get('new_value')
    if new_value is None:
        new_value = ''
    if not isinstance(new_value, (str, int, float, bool)):
        new_value = str(new_value)
    raw_df = raw_df.copy()
    raw_df.at[raw_df.index[row_index], column_name] = new_value
    raw_df = sanitize_dataframe(raw_df)
    data_store[data_id]['raw_df'] = raw_df
    return jsonify({'success': True, 'updated_value': _json_safe_value(raw_df.at[raw_df.index[row_index], column_name])})


@app.post('/api/add_raw_row')
def api_add_raw_row():
    """在原始数据末尾添加一行空行"""
    payload = request.json or {}
    data_id = payload.get('data_id')
    if not data_id or data_id not in data_store:
        return jsonify({'error': '数据不存在或已过期'}), 400
    raw_df = data_store[data_id].get('raw_df')
    if raw_df is None or not isinstance(raw_df, pd.DataFrame):
        return jsonify({'error': '没有原始数据'}), 400
    new_row = {col: '' for col in raw_df.columns}
    raw_df = pd.concat([raw_df, pd.DataFrame([new_row])], ignore_index=True)
    raw_df = sanitize_dataframe(raw_df)
    data_store[data_id]['raw_df'] = raw_df
    return jsonify({
        'success': True,
        'columns': [str(c) for c in raw_df.columns],
        'total_rows': int(len(raw_df)),
        'preview': dataframe_preview(raw_df)
    })


@app.post('/api/delete_raw_row')
def api_delete_raw_row():
    """删除原始数据中的指定行"""
    payload = request.json or {}
    data_id = payload.get('data_id')
    if not data_id or data_id not in data_store:
        return jsonify({'error': '数据不存在或已过期'}), 400
    raw_df = data_store[data_id].get('raw_df')
    if raw_df is None or not isinstance(raw_df, pd.DataFrame) or raw_df.empty:
        return jsonify({'error': '没有原始数据'}), 400
    try:
        row_index = int(payload.get('row_index'))
    except (TypeError, ValueError):
        return jsonify({'error': 'row_index 必须是整数'}), 400
    if row_index < 0 or row_index >= len(raw_df):
        return jsonify({'error': f'row_index 越界: {row_index}'}), 400
    raw_df = raw_df.drop(raw_df.index[row_index]).reset_index(drop=True)
    raw_df = sanitize_dataframe(raw_df)
    data_store[data_id]['raw_df'] = raw_df
    return jsonify({
        'success': True,
        'columns': [str(c) for c in raw_df.columns],
        'total_rows': int(len(raw_df)),
        'preview': dataframe_preview(raw_df)
    })


@app.post('/api/update_cell')
def api_update_cell():
    payload = request.json or {}
    data_id = payload.get('data_id')
    if not data_id or data_id not in data_store:
        return jsonify({'error': '数据不存在或已过期'}), 400
    tidy_df = data_store[data_id].get('tidy_df')
    if tidy_df is None or not isinstance(tidy_df, pd.DataFrame) or tidy_df.empty:
        return jsonify({'error': '没有可编辑的整理结果，请先执行整理'}), 400
    try:
        row_index = int(payload.get('row_index'))
    except (TypeError, ValueError):
        return jsonify({'error': 'row_index 必须是整数'}), 400
    column_name = str(payload.get('column_name') or '').strip()
    if not column_name:
        return jsonify({'error': 'column_name 不能为空'}), 400
    if column_name not in tidy_df.columns:
        return jsonify({'error': f'列不存在: {column_name}'}), 400
    if row_index < 0 or row_index >= len(tidy_df):
        return jsonify({'error': f'row_index 越界: {row_index}'}), 400
    new_value = payload.get('new_value')
    if new_value is None:
        new_value = ''
    if not isinstance(new_value, (str, int, float, bool)):
        new_value = str(new_value)
    tidy_df = tidy_df.copy()
    tidy_df.at[tidy_df.index[row_index], column_name] = new_value
    tidy_df = sanitize_dataframe(tidy_df)
    data_store[data_id]['tidy_df'] = tidy_df
    return jsonify({'success': True, 'updated_value': _json_safe_value(tidy_df.at[tidy_df.index[row_index], column_name])})


@app.post('/api/export')
def api_export():
    payload = request.json or {}
    data_id = payload.get('data_id')
    if not data_id or data_id not in data_store:
        return jsonify({'error': '数据不存在或已过期'}), 400
    tidy_df = data_store[data_id].get('tidy_df')
    if tidy_df is None or not isinstance(tidy_df, pd.DataFrame) or tidy_df.empty:
        return jsonify({'error': '没有可导出的整理结果，请先执行整理'}), 400
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sanitize_dataframe(tidy_df).to_excel(writer, index=False, sheet_name='TidyResult')
    output.seek(0)
    base_name = raw_file_store.get(data_id, {}).get('filename', 'tidy_result').rsplit('.', 1)[0]
    return send_file(output, as_attachment=True, download_name=f'{base_name}_tidy.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _open_browser() -> None:
    webbrowser.open(f'http://127.0.0.1:{PORT}')


if __name__ == '__main__':
    is_frozen = bool(getattr(sys, 'frozen', False))
    Timer(1.0, _open_browser).start()
    app.run(host='127.0.0.1', port=PORT, debug=False, use_reloader=not is_frozen)
