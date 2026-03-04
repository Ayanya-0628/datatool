"""
LLM 数据整理模块 v2 - 通过大模型 API 理解任意数据结构并生成转换代码
支持硅基流动 (SiliconFlow) + DeepSeek 模型

核心改进 v2：
- 大幅增强 System Prompt，明确处理多区域交错布局
- 增加数据读取范围（覆盖全文件）
- 错误重试时附带更多上下文
- 增加 max_retries 到 2
"""

import json
import re
import traceback
import openpyxl
import pandas as pd
from io import BytesIO

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False


# 默认配置
DEFAULT_API_BASE = "https://api.siliconflow.cn/v1"
DEFAULT_MODEL = "deepseek-ai/DeepSeek-V3"


def read_excel_raw(file_content, sheet_name=None, max_rows=80, max_cols=25):
    """
    用 openpyxl 读取 Excel 原始内容（含合并单元格），
    返回文本描述让 LLM 能理解数据结构。
    """
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    total_rows = ws.max_row or 0
    total_cols = ws.max_column or 0

    # 收集合并单元格信息
    merged_info = []
    merged_fill = {}

    for m in ws.merged_cells.ranges:
        val = ws.cell(m.min_row, m.min_col).value
        span_desc = f"R{m.min_row}C{m.min_col}:R{m.max_row}C{m.max_col}"
        merged_info.append(f"  {span_desc} => \"{val}\"  (跨{m.max_row-m.min_row+1}行{m.max_col-m.min_col+1}列)")
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                merged_fill[(r, c)] = val

    # 构建文本描述
    lines = []
    lines.append(f"=== Excel 工作表: {ws.title} ===")
    lines.append(f"总行数: {total_rows}, 总列数: {total_cols}")
    lines.append(f"")
    lines.append(f"合并单元格 ({len(merged_info)} 个):")
    for info in sorted(merged_info):
        lines.append(info)
    lines.append("")
    lines.append("全部单元格内容:")

    rows_to_show = min(max_rows, total_rows)
    for r in range(1, rows_to_show + 1):
        row_data = []
        for c in range(1, min(max_cols + 1, total_cols + 1)):
            val = ws.cell(r, c).value
            if val is None and (r, c) in merged_fill:
                val = merged_fill[(r, c)]
            if val is not None:
                sv = str(val)
                if len(sv) > 20:
                    sv = sv[:20] + ".."
                row_data.append(f"(R{r},C{c})={sv}")
        if row_data:
            lines.append(f"  Row{r}: {', '.join(row_data)}")
        else:
            lines.append(f"  Row{r}: (空行)")

    # 如果还有剩余行，继续输出
    if total_rows > rows_to_show:
        lines.append(f"\n... 省略 Row{rows_to_show+1}~Row{total_rows} ...")

    wb.close()
    return "\n".join(lines)


def call_llm(api_key, prompt, system_prompt=None,
             api_base=None, model=None):
    """调用 LLM API（OpenAI 兼容格式）"""
    if not HAS_REQUESTS:
        raise RuntimeError("缺少 requests 库")

    base = (api_base or DEFAULT_API_BASE).rstrip("/")
    mdl = model or DEFAULT_MODEL
    url = f"{base}/chat/completions"

    messages = []
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt})
    messages.append({"role": "user", "content": prompt})

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }

    payload = {
        "model": mdl,
        "messages": messages,
        "temperature": 0.05,
        "max_tokens": 8192
    }

    resp = requests.post(url, headers=headers, json=payload, timeout=180)
    resp.raise_for_status()
    result = resp.json()

    return result["choices"][0]["message"]["content"]


SYSTEM_PROMPT = """你是一个专业的数据整理助手。用户会给你一个 Excel 工作表的原始单元格数据。

## 你的任务

生成 Python 代码将这个复杂的 Excel 转换为**标准分析用宽表格式**。

## ⚠️ 绝对禁止写死行列号

由于文本丢失了部分空间直觉，**你绝对不能在代码中硬编码任何具体的行列数字**（如 `min_row=16`, `range(3, 15)` 等）。
你必须生成**动态搜索表块**的代码。

## 数据布局与动态探测策略

1. **结构特点**：一个表里散布着多个独立的数据表块。每个表块的**起点都是一个跨多列的标题合并单元格**（比如 "分蘖期干物质"、"孕穗-齐穗期积累" 等）。
2. **如何动态定位（请在代码中实现类似逻辑）**：
   - 第一步：遍历 `ws.merged_cells.ranges`，凡是值不为空（排除 None/处理/重复）、且位于某一行的起始位置、跨越多列的，都认为是"表块标题"。
   - 第二步：根据标题合并单元格的 `min_row` 和 `min_col`。它的正下方（如 `min_row+1`, `min_row+2`）通常是列名行（指标名如"茎(g)"、处理、重复等）。
   - 第三步：接着向下遍历获取该区域的独立数据，直到遇到空行跳出。
   - 第四步：每个表块独立提取出一份 DataFrame。

3. **左右交错问题**：不同表块（如干物质在左、积累量在右）存在于同一行中。靠着 `min_col` 与 `max_col` 范围来确保你只读取该块内的列。

## 宽表格式要求

- **每行** = 处理名 × 重复编号。不要把处理合并，每行代表一个生物学重复的唯一记录（例如处理A重复1、处理A重复2）。
- **列** = 包含所有时期和指标。列名必须带表块标题前缀，如 `分蘖期干物质_茎(g)`、`分蘖-孕穗期积累_值`。
- 移除所有名为"平均"、"均值"、"平均值"的列及其对应的数据。

## 必须遵守的代码约束

1. 手填合并单元格缓冲：
```python
merged_vals = {}
for m in ws.merged_cells.ranges:
    val = ws.cell(m.min_row, m.min_col).value
    for r in range(m.min_row, m.max_row+1):
        for c in range(m.min_col, m.max_col+1):
            merged_vals[(r,c)] = val

def get_val(r, c):
    v = ws.cell(r, c).value
    return v if v is not None else merged_vals.get((r,c))
```
2. 表块横向合并(Merge)：把所有临时表块的数据，用 `pd.merge(..., on=['处理','重复'], how='outer')` 合并。注意统一 `处理` 列转为 `str` 且去除空格。如果遇到缺少列名的空表头，请根据上下文推断或忽略。

## 返回格式

仅返回以下 JSON 对象（不要 ```json 包装，不要文字）。
{
  "description": "说明你对动态解析思路的规划",
  "tables_found": [{"name": "此处列举你探测到的表名,非写死行列"}],
  "code": "完整的 Python 代码（def transform(file_content, sheet_name=None): ...）"
}

- 只能用 openpyxl, pandas, io.BytesIO"""


def analyze_and_transform(file_content, api_key, sheet_name=None,
                          api_base=None, model=None, max_retries=3):
    """
    主入口：用 LLM 分析 Excel 结构并执行转换。
    如果生成的代码执行失败，会自动将错误反馈给 LLM 重试。

    Returns:
        dict: {
            'success': bool,
            'description': str,
            'tables_found': list,
            'result_df': DataFrame or None,
            'error': str or None,
            'llm_response': str
        }
    """
    # Step 1: 读取 Excel 原始内容（增大范围以覆盖全文件）
    raw_text = read_excel_raw(file_content, sheet_name, max_rows=80, max_cols=25)

    # Step 2: 构建初始 prompt
    prompt = f"""请分析以下 Excel 工作表的数据结构，并生成转换代码。

{raw_text}

请仔细观察：
1. 有哪些独立的数据表块？各自在什么行列范围？
2. 是否有不同列区域的表块共享同一行号（左右交错）？
3. 每个表块的指标列有多少组？
4. 请生成 transform 函数，将所有数据合并为一张宽表。

返回 JSON 对象。"""

    description = ''
    tables_found = []
    last_response = ''
    code = ''

    for attempt in range(max_retries + 1):
        try:
            print(f"[LLM Tidy] 第 {attempt + 1}/{max_retries + 1} 次调用 LLM...")

            # 调用 LLM
            llm_response = call_llm(api_key, prompt, SYSTEM_PROMPT, api_base, model)
            last_response = llm_response

            # 解析响应
            cleaned = llm_response.strip()
            if cleaned.startswith("```"):
                cleaned = re.sub(r'^```\w*\s*', '', cleaned)
                cleaned = re.sub(r'\s*```$', '', cleaned)

            parsed = json.loads(cleaned)
            description = parsed.get("description", "")
            tables_found = parsed.get("tables_found", [])
            code = parsed.get("code", "")

            if not code:
                return {
                    'success': False, 'description': description,
                    'tables_found': tables_found, 'result_df': None,
                    'error': 'LLM 未返回转换代码', 'llm_response': llm_response
                }

            # 执行代码
            namespace = {
                'openpyxl': openpyxl,
                'pd': pd,
                'BytesIO': BytesIO
            }
            exec(code, namespace)
            transform_fn = namespace.get('transform')

            if not transform_fn:
                raise RuntimeError('生成的代码中没有 transform 函数')

            result_df = transform_fn(file_content, sheet_name)

            if not isinstance(result_df, pd.DataFrame):
                raise TypeError(f'transform 返回类型不正确: {type(result_df)}')

            if len(result_df) == 0:
                raise ValueError('transform 返回空 DataFrame')

            # 基本质量检查
            n_cols = len(result_df.columns)
            n_rows = len(result_df)
            null_ratio = result_df.isnull().sum().sum() / (n_cols * n_rows) if n_cols * n_rows > 0 else 0

            if null_ratio > 0.5:
                raise ValueError(
                    f'结果中空值占比过高 ({null_ratio:.1%})，'
                    f'共 {n_rows}行×{n_cols}列，'
                    f'可能是不同表块被错误合并了。'
                    f'请确保每个独立表块按各自的列范围单独提取，'
                    f'再按处理名+重复编号合并。'
                )

            print(f"[LLM Tidy] 成功！{n_rows}行 × {n_cols}列, 空值率 {null_ratio:.1%}")

            return {
                'success': True, 'description': description,
                'tables_found': tables_found, 'result_df': result_df,
                'error': None, 'llm_response': llm_response
            }

        except json.JSONDecodeError as e:
            print(f"[LLM Tidy] JSON 解析失败: {e}")
            if attempt < max_retries:
                prompt = (
                    f"你上次返回的不是合法 JSON，解析错误: {str(e)}\n"
                    f"你的原始回复前100字符: {last_response[:100]}\n"
                    f"请严格遵循格式要求，只返回 JSON 对象，不要加 ```json 标记或其他文字。"
                )
                continue
            return {
                'success': False, 'description': description,
                'tables_found': tables_found, 'result_df': None,
                'error': f'JSON 解析失败: {str(e)}',
                'llm_response': last_response
            }

        except Exception as e:
            error_msg = str(e)
            tb = traceback.format_exc()
            print(f"[LLM Tidy] 第 {attempt + 1} 次失败: {error_msg}")

            if attempt < max_retries:
                # 将错误反馈给 LLM 要求修正
                prompt = f"""你上次生成的代码执行出错了，请根据以下错误信息修正。

错误信息: {error_msg}

错误追踪 (最后500字符):
{tb[-500:]}

你上次生成的代码:
```python
{code[:3000] if code else '(无)'}
```

关键提示:
- 如果是"空值占比过高"的错误，说明不同列区域的表块被混在了同一行，请确保每个表块按自己的列范围独立提取
- 如果是 KeyError 或 IndexError，检查行列索引是否正确
- 确保填充合并单元格后再读取数据

请返回修正后的完整 JSON 对象（包含 description、tables_found、code）。"""
                continue

            return {
                'success': False, 'description': description,
                'tables_found': tables_found, 'result_df': None,
                'error': f'执行失败 (已重试 {max_retries} 次): {error_msg}',
                'llm_response': last_response
            }
