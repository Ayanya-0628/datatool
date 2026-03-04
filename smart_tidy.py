"""
Robust smart tidy parser for complex interleaved Excel sheets.

This module is designed for layouts where multiple subtables are interleaved
both vertically and horizontally in a single worksheet.
"""

from __future__ import annotations

from io import BytesIO
from typing import Dict, List, Optional, Tuple
import re

import openpyxl
import pandas as pd

AVERAGE_KEYWORDS = ("平均", "均值", "平均值", "average", "avg", "mean")
HEADER_STOP_WORDS = {"处理", "重复", "平均", "均值", "平均值"}


class TableSpec(dict):
    """Small helper type to clarify spec fields."""


def _clean_text(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _build_merged_lookup(ws):
    merged_values = {}
    merged_spans = {}
    for merged in ws.merged_cells.ranges:
        value = ws.cell(merged.min_row, merged.min_col).value
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                merged_values[(row, col)] = value
                merged_spans[(row, col)] = (
                    merged.min_row,
                    merged.max_row,
                    merged.min_col,
                    merged.max_col,
                )
    return merged_values, merged_spans


def _is_average_label(text: Optional[str]) -> bool:
    if not text:
        return False
    lower = text.lower()
    return any(keyword in lower for keyword in AVERAGE_KEYWORDS)


def _ranges_overlap(a_start: int, a_end: int, b_start: int, b_end: int) -> bool:
    return a_start <= b_end and b_start <= a_end


def _to_digit(token: str) -> Optional[int]:
    if not token:
        return None
    match = re.search(r"\d+", token)
    if match:
        return int(match.group(0))

    mapping = {
        "一": 1,
        "二": 2,
        "三": 3,
        "四": 4,
        "五": 5,
        "六": 6,
        "七": 7,
        "八": 8,
        "九": 9,
        "十": 10,
    }
    return mapping.get(token)


def _parse_repeat_label(text: Optional[str]) -> Optional[str]:
    if not text:
        return None

    compact = re.sub(r"\s+", "", text)
    if _is_average_label(compact):
        return "__AVG__"

    match = re.search(r"重复([一二三四五六七八九十\d]+)", compact)
    if match:
        number = _to_digit(match.group(1))
        if number is not None:
            return f"重复{number}"

    if compact in {"1", "2", "3", "4", "5"}:
        return f"重复{compact}"

    return None


def _normalize_metric_name(metric: Optional[str], table_title: str) -> str:
    text = _clean_text(metric)
    if not text or text == "处理":
        return "值"
    if text == table_title:
        return "值"
    return text


def _coerce_value(value):
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _first_non_empty(values):
    for value in values:
        if value is None:
            continue
        if isinstance(value, float) and pd.isna(value):
            continue
        return value
    return None


def _pick_treatment_from_row(
    row: int,
    key_col_start: int,
    key_col_end: int,
    get_val,
) -> Optional[str]:
    key_values = []
    for col in range(key_col_start, key_col_end + 1):
        value = _clean_text(get_val(row, col))
        if value:
            key_values.append(value)

    if not key_values:
        return None

    candidate = key_values[-1]
    if candidate in HEADER_STOP_WORDS:
        return None
    if _is_average_label(candidate):
        return None
    if candidate.startswith("重复"):
        return None
    return candidate


def _is_title_candidate(value: Optional[str]) -> bool:
    text = _clean_text(value)
    if not text:
        return False
    if text in HEADER_STOP_WORDS:
        return False
    if _is_average_label(text):
        return False

    metric_prefixes = ("茎", "叶", "穗", "总重")
    if text.startswith(metric_prefixes):
        return False

    return True


def _find_key_anchor_for_title(ws, get_val, merged_spans, merge_range) -> Optional[Tuple[int, int, int]]:
    title_row = merge_range.min_row
    col_start = max(1, merge_range.min_col - 2)
    col_end = merge_range.max_col

    for row in range(title_row, min(ws.max_row, title_row + 2) + 1):
        for col in range(col_start, col_end + 1):
            value = _clean_text(get_val(row, col))
            if value != "处理":
                continue

            span = merged_spans.get((row, col))
            if span:
                _, _, key_start, key_end = span
            else:
                key_start = col
                key_end = col
            return row, key_start, key_end

    return None


def _expand_value_col_end(ws, key_row: int, key_col_end: int, current_end: int, get_val) -> int:
    """
    Expand value-column boundary using merged metric headers near key/header rows.
    This is required for cases where a title merge is narrower than the real metric area.
    """
    expanded_end = current_end
    changed = True

    while changed:
        changed = False
        for merged in ws.merged_cells.ranges:
            row_span = merged.max_row - merged.min_row + 1
            col_span = merged.max_col - merged.min_col + 1
            if row_span != 1 or col_span < 2:
                continue
            if not (key_row <= merged.min_row <= key_row + 1):
                continue
            if merged.min_col <= key_col_end:
                continue
            if merged.min_col > expanded_end + 2:
                continue

            value = _clean_text(get_val(merged.min_row, merged.min_col))
            if not value or value in HEADER_STOP_WORDS or _is_average_label(value):
                continue

            if merged.max_col > expanded_end:
                expanded_end = merged.max_col
                changed = True

    return expanded_end


def _detect_table_specs(ws, get_val, merged_spans) -> List[TableSpec]:
    specs: List[TableSpec] = []

    merged_ranges = sorted(
        ws.merged_cells.ranges,
        key=lambda item: (item.min_row, item.min_col, item.max_col),
    )

    for merged in merged_ranges:
        row_span = merged.max_row - merged.min_row + 1
        col_span = merged.max_col - merged.min_col + 1
        if row_span != 1 or col_span < 4:
            continue

        title = _clean_text(get_val(merged.min_row, merged.min_col))
        if not _is_title_candidate(title):
            continue

        key_anchor = _find_key_anchor_for_title(ws, get_val, merged_spans, merged)
        if key_anchor is None:
            continue

        key_row, key_col_start, key_col_end = key_anchor
        value_col_start = max(key_col_end + 1, merged.min_col)
        value_col_end = _expand_value_col_end(
            ws=ws,
            key_row=key_row,
            key_col_end=key_col_end,
            current_end=merged.max_col,
            get_val=get_val,
        )

        if value_col_start > value_col_end:
            continue

        specs.append(
            TableSpec(
                title=title,
                title_row=merged.min_row,
                key_row=key_row,
                key_col_start=key_col_start,
                key_col_end=key_col_end,
                value_col_start=value_col_start,
                value_col_end=value_col_end,
                table_col_start=min(key_col_start, merged.min_col),
                table_col_end=max(value_col_end, key_col_end),
            )
        )

    deduped: List[TableSpec] = []
    seen = set()
    for spec in specs:
        key = (spec["title_row"], spec["table_col_start"], spec["table_col_end"], spec["title"])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(spec)

    deduped.sort(key=lambda item: (item["title_row"], item["table_col_start"]))

    for index, spec in enumerate(deduped):
        boundary_end = ws.max_row
        for candidate in deduped[index + 1 :]:
            if not _ranges_overlap(
                spec["table_col_start"],
                spec["table_col_end"],
                candidate["table_col_start"],
                candidate["table_col_end"],
            ):
                continue
            boundary_end = candidate["title_row"] - 1
            break

        end_row = _find_table_end_row(spec, boundary_end, get_val)
        spec["end_row"] = max(end_row, spec["key_row"] + 1)

    return deduped


def _find_table_end_row(spec: TableSpec, boundary_end: int, get_val) -> int:
    start_row = spec["key_row"] + 1
    last_data_row = start_row - 1
    empty_streak = 0

    for row in range(start_row, boundary_end + 1):
        treatment = _pick_treatment_from_row(
            row,
            spec["key_col_start"],
            spec["key_col_end"],
            get_val,
        )
        if treatment:
            last_data_row = row
            empty_streak = 0
        else:
            if last_data_row >= start_row:
                empty_streak += 1
                if empty_streak >= 2:
                    break

    return max(last_data_row, start_row)


def _extract_subtable_dataframe(ws, get_val, spec: TableSpec, drop_avg: bool) -> pd.DataFrame:
    key_row = spec["key_row"]
    sub_header_row = key_row + 1 if key_row + 1 <= spec["end_row"] else key_row

    column_defs = []
    has_repeat_sub_headers = False

    for col in range(spec["value_col_start"], spec["value_col_end"] + 1):
        top_label = _clean_text(get_val(key_row, col))
        sub_label = _clean_text(get_val(sub_header_row, col))
        repeat_label = _parse_repeat_label(sub_label)
        if repeat_label is not None:
            has_repeat_sub_headers = True

        column_defs.append(
            {
                "col": col,
                "metric": _normalize_metric_name(top_label, spec["title"]),
                "repeat": repeat_label,
                "sub": sub_label,
            }
        )

    use_two_header_rows = has_repeat_sub_headers and sub_header_row > key_row
    data_start_row = key_row + (2 if use_two_header_rows else 1)

    repeat_metric_cols: Dict[str, Dict[str, List[int]]] = {}
    for definition in column_defs:
        repeat_label = definition["repeat"]
        if repeat_label is None:
            if use_two_header_rows:
                continue
            repeat_label = "重复1"

        if drop_avg and repeat_label == "__AVG__":
            continue

        metric_name = definition["metric"]
        if drop_avg and _is_average_label(metric_name):
            continue

        repeat_metric_cols.setdefault(repeat_label, {}).setdefault(metric_name, []).append(definition["col"])

    records = []
    for row in range(data_start_row, spec["end_row"] + 1):
        treatment = _pick_treatment_from_row(
            row,
            spec["key_col_start"],
            spec["key_col_end"],
            get_val,
        )
        if not treatment:
            continue

        for repeat_label, metric_map in repeat_metric_cols.items():
            if repeat_label == "__AVG__":
                continue

            record = {"处理": treatment, "重复": repeat_label}
            non_null_values = 0

            for metric_name, cols in metric_map.items():
                value = _first_non_empty([_coerce_value(get_val(row, col)) for col in cols])
                if value is not None:
                    non_null_values += 1
                record[f"{spec['title']}_{metric_name}"] = value

            if non_null_values > 0:
                records.append(record)

    if not records:
        return pd.DataFrame(columns=["处理", "重复"])

    df = pd.DataFrame(records)
    df = _coalesce_duplicate_rows(df)
    return df


def _coalesce_duplicate_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    key_columns = ["处理", "重复"]
    metric_columns = [col for col in df.columns if col not in key_columns]

    if not metric_columns:
        return df.drop_duplicates(subset=key_columns).reset_index(drop=True)

    def pick_first_valid(series):
        for value in series:
            if value is None:
                continue
            if isinstance(value, float) and pd.isna(value):
                continue
            return value
        return None

    grouped = (
        df.groupby(key_columns, sort=False, as_index=False)[metric_columns]
        .agg(pick_first_valid)
        .reset_index(drop=True)
    )
    return grouped


def _merge_subtable_dataframes(frames: List[pd.DataFrame]) -> pd.DataFrame:
    valid_frames = [frame for frame in frames if frame is not None and not frame.empty]
    if not valid_frames:
        return pd.DataFrame(columns=["处理", "重复"])

    merged = valid_frames[0].copy()
    for frame in valid_frames[1:]:
        merged = pd.merge(merged, frame, on=["处理", "重复"], how="outer")

    merged["处理"] = merged["处理"].astype(str).str.strip()
    merged["重复"] = merged["重复"].astype(str).str.strip()

    ordered_cols = ["处理", "重复"] + [
        col for col in merged.columns if col not in {"处理", "重复"}
    ]
    merged = merged[ordered_cols]

    merged = merged.drop_duplicates(subset=["处理", "重复"]).reset_index(drop=True)
    return merged


def _serialize_preview(df: pd.DataFrame, preview_rows: int = 20) -> List[Dict[str, object]]:
    if df.empty:
        return []

    preview_df = df.head(preview_rows).copy()
    preview_df = preview_df.where(pd.notna(preview_df), None)

    records = []
    for row in preview_df.to_dict(orient="records"):
        record = {}
        for key, value in row.items():
            if value is not None and not isinstance(value, (str, int, float, bool)):
                record[key] = str(value)
            else:
                record[key] = value
        records.append(record)
    return records


def _parse_excel(file_content: bytes, sheet_name: Optional[str], drop_avg: bool):
    workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    try:
        if sheet_name and sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
        else:
            ws = workbook.active

        merged_values, merged_spans = _build_merged_lookup(ws)

        def get_val(row: int, col: int):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                return value
            return merged_values.get((row, col))

        specs = _detect_table_specs(ws, get_val, merged_spans)
        subtable_frames = []
        summaries = []

        for spec in specs:
            frame = _extract_subtable_dataframe(ws, get_val, spec, drop_avg=drop_avg)
            if frame is None or frame.empty:
                continue

            subtable_frames.append(frame)
            summaries.append(
                {
                    "title": spec["title"],
                    "rows": int(len(frame)),
                    "columns": int(len(frame.columns) - 2),
                    "row_range": [int(spec["title_row"]), int(spec["end_row"])],
                    "column_range": [int(spec["table_col_start"]), int(spec["table_col_end"])],
                    "factor_headers": ["处理", "重复"],
                }
            )

        merged_df = _merge_subtable_dataframes(subtable_frames)

        return {
            "sheet_name": ws.title,
            "merged_count": len(ws.merged_cells.ranges),
            "specs": specs,
            "summaries": summaries,
            "table_frames": subtable_frames,
            "merged_df": merged_df,
        }
    finally:
        workbook.close()


def scan_excel_structure(file_content, sheet_name=None):
    """Scan workbook and build a robust multi-subtable parsing plan."""
    if not file_content:
        return {"error": "Excel 内容为空"}

    parsed = _parse_excel(file_content, sheet_name, drop_avg=True)
    merged_df = parsed["merged_df"]

    if merged_df.empty:
        return {"error": "未检测到有效的子表数据"}

    title_parts = [item["title"] for item in parsed["summaries"] if item.get("title")]
    title = " / ".join(title_parts) if title_parts else parsed["sheet_name"]

    return {
        "title": title,
        "sheet_name": parsed["sheet_name"],
        "merged_count": parsed["merged_count"],
        "sub_table_count": len(parsed["summaries"]),
        "sub_tables": parsed["summaries"],
        "total_rows": int(len(merged_df)),
        "total_cols": int(len(merged_df.columns)),
        "preview": _serialize_preview(merged_df),
        "_file_content": file_content,
        "_sheet_name": sheet_name,
        "_drop_avg_df": merged_df,
        "_spec_count": len(parsed["specs"]),
    }


def execute_smart_tidy(scan_result, options=None):
    """Execute transformation using scan result and return wide-format DataFrame."""
    if options is None:
        options = {}

    drop_avg = bool(options.get("drop_avg", True))

    if drop_avg and isinstance(scan_result.get("_drop_avg_df"), pd.DataFrame):
        return scan_result["_drop_avg_df"].copy()

    file_content = scan_result.get("_file_content")
    sheet_name = scan_result.get("_sheet_name")
    if not file_content:
        return pd.DataFrame(columns=["处理", "重复"])

    parsed = _parse_excel(file_content, sheet_name, drop_avg=drop_avg)
    return parsed["merged_df"]
